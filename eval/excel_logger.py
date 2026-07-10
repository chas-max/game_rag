"""统一 Excel 指标记录 - 追加式,两个 sheet,带版本标记列。

- `runs` sheet: 每次评测一行(汇总指标)
- `per_query` sheet: 每条问题一行(明细,便于分析)
两 sheet 均含版本标记列(run_id/version_label/run_time/git_commit/...),
不同项目版本(基线 vs 优化后)的指标以追加行形式保留,便于横向对比。
"""

from __future__ import annotations

import json
import os
from typing import Any

from openpyxl import Workbook, load_workbook

# ── 列定义(顺序即 Excel 列顺序) ─────────────────────────────────────

# runs sheet: 版本标记 + 核心指标
RUNS_COLUMNS = [
    # 版本标记
    "run_id", "version_label", "run_time", "git_commit", "git_dirty",
    "benchmark_id", "config_snapshot", "note", "with_generation",
    # 规模
    "n_queries",
    # 5 核心指标
    "accuracy", "recall", "throughput", "faithfulness", "relevance",
]

# per_query sheet: 版本标记子集 + 单条问题明细
PER_QUERY_COLUMNS = [
    "run_id", "version_label", "run_time", "git_commit", "with_generation", "benchmark_id",
    "query_id", "game_name", "question",
    "accuracy", "recall", "faithfulness", "relevance", "latency_ms",
]


def _cell(v: Any) -> Any:
    """把值转为 Excel 友好类型: dict/list 转 JSON 字符串,其余原样。"""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return v


def _load_or_create(path: str):
    """打开已有工作簿,否则新建(去掉默认 Sheet)。文件被占用时给清晰报错。"""
    if os.path.exists(path):
        try:
            return load_workbook(path)
        except PermissionError:
            raise SystemExit(f"[excel] {path} 被占用(可能在 Excel 中打开),请关闭后重试。")
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def _ensure_sheet(wb, name: str, columns: list[str]):
    """获取 sheet;若为空则写入表头并设置列宽。"""
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
    if ws.cell(1, 1).value is None:
        # 新 sheet 的 max_row 默认为 1 但单元格为空,直接把表头写到第 1 行,
        # 避免用 append() 导致首行空置、表头错位(进而重载时被重复写表头)。
        for i, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=i, value=col)
            width = min(max(len(col) * 2 + 4, 12), 50)
            ws.column_dimensions[cell.column_letter].width = width
    return ws


def log_run(
    out_path: str,
    run_meta: dict,
    summary: dict,
    per_query_rows: list[dict],
) -> None:
    """把一次评测的汇总与明细追加写入 Excel(两个 sheet)。

    Args:
        out_path: Excel 路径(不存在则新建,存在则追加)。
        run_meta: 版本标记字段(run_id/version_label/run_time/git_commit/git_dirty/
            benchmark_id/config_snapshot/note/with_generation)。
        summary: 汇总指标(n_queries/recall@topk/...);缺字段留空。
        per_query_rows: 每条问题的明细指标列表。
    """
    wb = _load_or_create(out_path)

    # runs sheet
    ws_runs = _ensure_sheet(wb, "runs", RUNS_COLUMNS)
    runs_row = {**run_meta, **summary}
    ws_runs.append([_cell(runs_row.get(c)) for c in RUNS_COLUMNS])

    # per_query sheet
    ws_pq = _ensure_sheet(wb, "per_query", PER_QUERY_COLUMNS)
    pq_meta_keys = ["run_id", "version_label", "run_time", "git_commit", "with_generation", "benchmark_id"]
    pq_meta = {k: run_meta.get(k) for k in pq_meta_keys}
    for r in per_query_rows:
        row = {**pq_meta, **r}
        ws_pq.append([_cell(row.get(c)) for c in PER_QUERY_COLUMNS])

    try:
        wb.save(out_path)
    except PermissionError:
        raise SystemExit(f"[excel] {out_path} 被占用(可能在 Excel 中打开),请关闭后重试。")
